"""
A.L.E.C. Excel Engine — Read, write, edit, and export Excel files.
Supports .xlsx, .xls, .csv, and .numbers (read-only via csv export).

This gives A.L.E.C. the ability to:
1. Read spreadsheets and understand their contents
2. Generate reports and export data as Excel files
3. Edit existing spreadsheets (add rows, update cells, add sheets)
4. Convert between formats (CSV ↔ Excel)
5. Analyze spreadsheet data and generate insights
"""

import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

logger = logging.getLogger("alec.excel")

EXPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "exports"


class ExcelEngine:
    """Read, write, and manipulate Excel files."""

    def __init__(self):
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        self._check_deps()

    def _check_deps(self):
        """Check if pandas and openpyxl are available."""
        self.has_pandas = False
        self.has_openpyxl = False
        try:
            import pandas
            self.has_pandas = True
        except ImportError:
            logger.warning("pandas not installed — Excel features limited")
        try:
            import openpyxl
            self.has_openpyxl = True
        except ImportError:
            logger.warning("openpyxl not installed — .xlsx write disabled")

    # ── Read ─────────────────────────────────────────────────────

    def read_file(self, filepath: str, sheet_name: Optional[str] = None,
                  max_rows: int = 1000) -> dict:
        """
        Read a spreadsheet file and return its contents.
        Supports: .xlsx, .xls, .csv, .tsv
        Returns: {sheets: [{name, headers, rows, row_count, col_count}]}
        """
        if not self.has_pandas:
            return {"error": "pandas not installed. Run: pip install pandas openpyxl"}

        import pandas as pd
        fpath = Path(filepath)

        if not fpath.exists():
            return {"error": f"File not found: {filepath}"}

        result = {"filepath": filepath, "sheets": []}

        try:
            suffix = fpath.suffix.lower()

            if suffix in (".csv", ".tsv"):
                sep = "\t" if suffix == ".tsv" else ","
                df = pd.read_csv(fpath, sep=sep, nrows=max_rows)
                result["sheets"].append(self._df_to_dict(df, "Sheet1"))

            elif suffix in (".xlsx", ".xls"):
                xls = pd.ExcelFile(fpath)
                sheet_names = [sheet_name] if sheet_name else xls.sheet_names
                for sn in sheet_names:
                    if sn in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sn, nrows=max_rows)
                        result["sheets"].append(self._df_to_dict(df, sn))
            else:
                return {"error": f"Unsupported format: {suffix}"}

        except Exception as e:
            return {"error": f"Failed to read file: {e}"}

        return result

    def _df_to_dict(self, df, sheet_name: str) -> dict:
        """Convert a pandas DataFrame to a serializable dict."""
        import pandas as pd
        # Convert to native Python types for JSON serialization
        rows = []
        for _, row in df.iterrows():
            rows.append({
                str(col): (None if pd.isna(val) else
                           val.isoformat() if hasattr(val, 'isoformat') else
                           str(val) if not isinstance(val, (int, float, bool, str)) else val)
                for col, val in row.items()
            })

        return {
            "name": sheet_name,
            "headers": [str(c) for c in df.columns.tolist()],
            "rows": rows,
            "row_count": len(df),
            "col_count": len(df.columns),
        }

    # ── Write / Export ───────────────────────────────────────────

    def export_to_excel(self, data: dict, filename: str = None) -> dict:
        """
        Export data to an Excel file.

        data format:
        {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["Name", "Value", "Date"],
                    "rows": [
                        {"Name": "Item 1", "Value": 100, "Date": "2024-01-01"},
                        ...
                    ]
                }
            ]
        }

        Returns: {filepath, filename, size_bytes}
        """
        if not self.has_pandas or not self.has_openpyxl:
            return {"error": "pandas and openpyxl required. Run: pip install pandas openpyxl"}

        import pandas as pd

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"alec_export_{timestamp}.xlsx"

        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        filepath = EXPORTS_DIR / filename

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                for sheet in data.get("sheets", []):
                    df = pd.DataFrame(sheet.get("rows", []))
                    if sheet.get("headers"):
                        df = df[sheet["headers"]] if all(h in df.columns for h in sheet["headers"]) else df
                    sheet_name = sheet.get("name", "Sheet1")[:31]  # Excel limit
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            size = filepath.stat().st_size
            logger.info(f"Exported Excel: {filepath} ({size} bytes)")
            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "size_bytes": size,
                "download_url": f"/exports/{filename}",
            }

        except Exception as e:
            return {"error": f"Export failed: {e}"}

    def export_to_csv(self, data: dict, filename: str = None) -> dict:
        """Export data to CSV format."""
        if not self.has_pandas:
            return {"error": "pandas required"}

        import pandas as pd

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"alec_export_{timestamp}.csv"

        filepath = EXPORTS_DIR / filename

        try:
            sheet = data.get("sheets", [{}])[0]
            df = pd.DataFrame(sheet.get("rows", []))
            df.to_csv(filepath, index=False)

            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "size_bytes": filepath.stat().st_size,
            }
        except Exception as e:
            return {"error": f"CSV export failed: {e}"}

    # ── Edit ─────────────────────────────────────────────────────

    def edit_file(self, filepath: str, operations: list[dict]) -> dict:
        """
        Edit an existing Excel file.

        operations: [
            {"action": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": "Hello"},
            {"action": "add_row", "sheet": "Sheet1", "row": {"Name": "New", "Value": 42}},
            {"action": "add_sheet", "name": "NewSheet"},
            {"action": "delete_row", "sheet": "Sheet1", "row_index": 5},
        ]
        """
        if not self.has_openpyxl:
            return {"error": "openpyxl required"}

        import openpyxl

        fpath = Path(filepath)
        if not fpath.exists():
            return {"error": f"File not found: {filepath}"}

        try:
            wb = openpyxl.load_workbook(filepath)
            changes = 0

            for op in operations:
                action = op.get("action")

                if action == "set_cell":
                    sheet = wb[op["sheet"]] if op.get("sheet") in wb.sheetnames else wb.active
                    sheet[op["cell"]] = op["value"]
                    changes += 1

                elif action == "add_row":
                    sheet = wb[op["sheet"]] if op.get("sheet") in wb.sheetnames else wb.active
                    row_data = op.get("row", {})
                    # Find headers in first row
                    headers = [cell.value for cell in sheet[1]]
                    new_row = [row_data.get(h) for h in headers]
                    sheet.append(new_row)
                    changes += 1

                elif action == "add_sheet":
                    name = op.get("name", f"Sheet{len(wb.sheetnames) + 1}")
                    wb.create_sheet(title=name[:31])
                    changes += 1

                elif action == "delete_row":
                    sheet = wb[op["sheet"]] if op.get("sheet") in wb.sheetnames else wb.active
                    idx = op.get("row_index", 1)
                    sheet.delete_rows(idx)
                    changes += 1

            wb.save(filepath)
            return {"success": True, "changes": changes, "filepath": filepath}

        except Exception as e:
            return {"error": f"Edit failed: {e}"}

    # ── Analyze ──────────────────────────────────────────────────

    def analyze(self, filepath: str) -> dict:
        """Quick analysis of a spreadsheet — summary stats, data types, etc."""
        if not self.has_pandas:
            return {"error": "pandas required"}

        import pandas as pd

        data = self.read_file(filepath)
        if "error" in data:
            return data

        analysis = {"filepath": filepath, "sheets": []}

        for sheet in data["sheets"]:
            df = pd.DataFrame(sheet["rows"])
            sheet_analysis = {
                "name": sheet["name"],
                "row_count": sheet["row_count"],
                "col_count": sheet["col_count"],
                "headers": sheet["headers"],
                "dtypes": {col: str(df[col].dtype) for col in df.columns},
                "null_counts": {col: int(df[col].isna().sum()) for col in df.columns},
            }

            # Numeric columns get stats
            numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
            if numeric_cols:
                stats = df[numeric_cols].describe().to_dict()
                sheet_analysis["numeric_stats"] = {
                    col: {k: round(v, 2) if isinstance(v, float) else v
                          for k, v in stats[col].items()}
                    for col in numeric_cols
                }

            analysis["sheets"].append(sheet_analysis)

        return analysis

    def get_status(self) -> dict:
        """Return Excel engine capabilities."""
        exports = list(EXPORTS_DIR.glob("*.xlsx")) + list(EXPORTS_DIR.glob("*.csv"))
        return {
            "pandas_available": self.has_pandas,
            "openpyxl_available": self.has_openpyxl,
            "exports_dir": str(EXPORTS_DIR),
            "export_count": len(exports),
            "supported_formats": {
                "read": [".xlsx", ".xls", ".csv", ".tsv"],
                "write": [".xlsx", ".csv"] if self.has_openpyxl else [".csv"],
                "edit": [".xlsx"] if self.has_openpyxl else [],
            },
        }
